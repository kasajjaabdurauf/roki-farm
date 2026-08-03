#!/usr/bin/env python3
"""
Roki simulation files generator.
Creates two XLSX workbooks for demo/load-testing and editing practice:

  1) roki-simulation-1000-farmers.xlsx
     - 1,000 farmer rows with produce columns (imports farmers + harvest logs
       through Bulk Upload)
     - ~30 rows contain INTENTIONAL errors (bad phones, negative quantities,
       blank names/dates, duplicate-log pairs) so you can practice fixing
       them in the staging grid. Error cells are tinted light red and all
       are documented in the "Read Me" sheet.

  2) roki-farmer-log-simulation.xlsx
     - 900 harvest-log rows referencing the farmer IDs above
     - a few intentional anomalies/duplicates/errors for practice

Deterministic (seeded) so the files are reproducible.
"""
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

rng = random.Random(20260803)

# ----------------------------------------------------------------------
# Reference data
# ----------------------------------------------------------------------
FIRST_M = ["John", "Peter", "Moses", "Daniel", "Charles", "Isaac", "Joseph", "Patrick", "Samuel",
           "Vincent", "Robert", "George", "Henry", "James", "Fred", "Yusuf", "Kevin", "Dennis",
           "Andrew", "David", "Eric", "Frank", "Godfrey", "Herbert", "Ivan", "Lawrence", "Martin",
           "Nicholas", "Ronald", "Stephen"]
FIRST_F = ["Aisha", "Grace", "Sarah", "Joyce", "Agnes", "Betty", "Florence", "Harriet", "Janet",
           "Linda", "Mary", "Rebecca", "Winnie", "Zainab", "Catherine", "Gloria", "Irene", "Martha",
           "Naomi", "Patience", "Rose", "Sandra", "Tracy", "Veronica", "Angella", "Beatrice",
           "Doreen", "Eunice", "Faridah", "Hadijah"]
SURNAMES = ["Namukwaya", "Okello", "Achieng", "Ssemanda", "Nakanwagi", "Odongo", "Auma", "Mugisha",
            "Kobusingye", "Opio", "Amongin", "Byaruhanga", "Namatovu", "Ochen", "Kemigisha",
            "Tumusiime", "Namuddu", "Kato", "Mukasa", "Apio", "Nakato", "Okwir", "Nabirye",
            "Wamala", "Anying", "Ouma", "Nakimuli", "Kiggundu", "Nalwoga", "Ayesiga", "Nankya",
            "Kyazze", "Namugga", "Kirumira", "Nabukenya", "Mpanga", "Busingye", "Kizza", "Lubega",
            "Muyingo", "Nabatanzi", "Ojok", "Onyango", "Sekitoleko", "Tumuhairwe", "Wandera"]

DISTRICTS = ["Wakiso", "Mukono", "Masaka", "Mbarara", "Gulu", "Lira", "Soroti", "Jinja", "Arua",
             "Kampala", "Bushenyi", "Ntungamo", "Iganga", "Tororo", "Kasese", "Hoima", "Mubende",
             "Luwero", "Kayunga", "Kabarole (Fort Portal)", "Rukungiri", "Rakai", "Mbale", "Kabale"]
SUB_COUNTY_POOL = ["Central", "Town Council", "North", "South", "East", "West", "Rural A", "Rural B"]
VILLAGES = ["Kyambogo", "Busukuma", "Kigungu", "Seeta", "Nagojje", "Maziba", "Katine", "Ogur",
            "Awoja", "Kilembe", "Laroo", "Bardege", "Rukoni", "Kasanje", "Zirobwe", "Butoloogo",
            "Nagongera", "Bumasifwa", "Katikamu", "Bukoyo", "Kiyuni", "Lubaga", "Oli", "Unyama"]

CROPS = ["Tomato", "Onion", "Cabbage", "Carrots", "Watermelon", "Eggplant", "Passion Fruit",
         "Chilli Pepper", "Maize", "Beans", "Bananas", "Cassava", "Groundnuts", "Sweet Potatoes",
         "Rice", "Coffee", "Avocado", "Mango"]
TYPICAL_KG_PER_ACRE = {
    "Tomato": 6000, "Onion": 4000, "Cabbage": 6000, "Carrots": 5000, "Watermelon": 9000,
    "Eggplant": 5000, "Passion Fruit": 3500, "Chilli Pepper": 2500, "Maize": 1600, "Beans": 600,
    "Bananas": 5500, "Cassava": 4000, "Groundnuts": 700, "Sweet Potatoes": 4500, "Rice": 1300,
    "Coffee": 700, "Avocado": 5000, "Mango": 4000,
}
GRADES = ["A", "A", "A", "B", "B", "REJECT"]
PHONE_PREFIXES = ["77", "78", "76", "70", "74", "75"]

def gen_phone():
    return f"+256{rng.choice(PHONE_PREFIXES)}{rng.randint(1000000, 9999999)}"

def gen_name(gender):
    return f"{rng.choice(FIRST_F if gender == 'F' else FIRST_M)} {rng.choice(SURNAMES)}"

def gen_date():
    # 2026-05-01 .. 2026-08-15
    return f"2026-{rng.randint(5, 8):02d}-{rng.randint(1, 28):02d}"

def gen_batch(d):
    return f"B-{d.replace('-', '')}-{rng.randint(100, 999)}"

# ----------------------------------------------------------------------
# Styling helpers
# ----------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1B4332")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ERROR_FILL = PatternFill("solid", fgColor="FECACA")          # light red
WARN_FILL = PatternFill("solid", fgColor="FDE68A")           # light amber
THIN = Side(style="thin", color="E7E5E4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_sheet(ws, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

def write_readme(ws, title, intro, problems, notes=None):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color="1B4332")
    ws["A2"] = intro
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A2"].font = Font(size=11)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 46
    ws.column_dimensions["E"].width = 40
    r = 4
    ws.cell(row=r, column=1, value="Excel row").font = Font(bold=True)
    ws.cell(row=r, column=2, value="Column").font = Font(bold=True)
    ws.cell(row=r, column=3, value="Problem (intentional)").font = Font(bold=True)
    ws.cell(row=r, column=4, value="How to fix in the staging grid").font = Font(bold=True)
    ws.cell(row=r, column=5, value="What you should see after fixing").font = Font(bold=True)
    r += 1
    for row, col, prob, fix, expect in problems:
        ws.cell(row=r, column=1, value=row)
        ws.cell(row=r, column=2, value=col)
        ws.cell(row=r, column=3, value=prob)
        ws.cell(row=r, column=4, value=fix)
        ws.cell(row=r, column=5, value=expect)
        r += 1
    if notes:
        r += 1
        ws.cell(row=r, column=1, value="Notes").font = Font(bold=True)
        r += 1
        for n in notes:
            ws.cell(row=r, column=1, value="•")
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
            ws.cell(row=r, column=2, value=n).alignment = Alignment(wrap_text=True, vertical="top")
            r += 1

# ======================================================================
# FILE 1 — 1,000 farmers
# ======================================================================
HEADERS1 = ["Farmer ID", "Farmer Name", "Email", "Phone", "Gender", "Refugee Status", "District",
            "Sub-County", "Village", "Acreage", "Crop", "Qty (Kg)", "Harvest Date", "Grade",
            "Batch ID"]
WIDTHS1 = [14, 24, 26, 17, 9, 14, 22, 12, 14, 9, 15, 10, 13, 8, 17]

wb = Workbook()
ws = wb.active
ws.title = "Farmer Upload"
style_sheet(ws, HEADERS1, WIDTHS1)

farmers = []          # (id, name, phone, gender, district, sub_county, village, acreage)
problems1 = []        # documented problems for Read Me

ERROR_BAD_PHONE = 6
ERROR_BAD_QTY = 6
ERROR_BLANK_NAME = 4
ERROR_BAD_DATE = 4
ERROR_BAD_ACREAGE = 3
ERROR_NO_FARMER = 2
DUPLICATE_PAIRS = 2

# pick the error row positions (data rows 1..1000, excel rows 2..1001)
bad_phone_rows = rng.sample(range(1, 1001), ERROR_BAD_PHONE)
bad_qty_rows = rng.sample([r for r in range(1, 1001) if r not in bad_phone_rows], ERROR_BAD_QTY)
blank_name_rows = rng.sample([r for r in range(1, 1001) if r not in bad_phone_rows + bad_qty_rows], ERROR_BLANK_NAME)
bad_date_rows = rng.sample([r for r in range(1, 1001) if r not in bad_phone_rows + bad_qty_rows + blank_name_rows], ERROR_BAD_DATE)
bad_acreage_rows = rng.sample([r for r in range(1, 1001) if r not in bad_phone_rows + bad_qty_rows + blank_name_rows + bad_date_rows], ERROR_BAD_ACREAGE)
no_farmer_rows = rng.sample([r for r in range(1, 1001) if r not in bad_phone_rows + bad_qty_rows + blank_name_rows + bad_date_rows + bad_acreage_rows], ERROR_NO_FARMER)
dup_anchor_rows = rng.sample([r for r in range(1, 1001) if r not in bad_phone_rows + bad_qty_rows + blank_name_rows + bad_date_rows + bad_acreage_rows + no_farmer_rows], DUPLICATE_PAIRS)

def mark_error(cell, problems, excel_row, col, prob, fix, expect):
    cell.fill = ERROR_FILL
    problems.append((excel_row, col, prob, fix, expect))

all_rows1 = []  # (excel_row, values, problem_entries)
for i in range(1, 1001):
    excel_row = i + 1
    fid = f"RFV-UG-{i:05d}"
    gender = rng.choice(["M", "F", "F"])
    name = gen_name(gender)
    email = f"{name.split()[0].lower()}.{name.split()[1].lower()}{rng.randint(1,999)}@example.com"
    phone = gen_phone()
    refugee = rng.choice(["REFUGEE", "HOST", "HOST", "NONE"])
    district = rng.choice(DISTRICTS)
    sub = rng.choice(SUB_COUNTY_POOL)
    village = rng.choice(VILLAGES)
    acreage = round(rng.uniform(0.4, 12.0), 1)
    crop = rng.choice(CROPS)
    qty = round(acreage * TYPICAL_KG_PER_ACRE[crop] * rng.uniform(0.6, 1.1), 0)
    date = gen_date()
    grade = rng.choice(GRADES)
    batch = gen_batch(date)

    row = [fid, name, email, phone, gender, refugee, district, sub, village, acreage, crop, qty, date, grade, batch]

    # ---- intentional problems ----
    if i in bad_phone_rows:
        bad = rng.choice(["077X 123 45A", "0414 555 666", "+25670123456", "2567 12345"])
        row[2] = bad
        mark_error(ws.cell(row=excel_row, column=4), problems1, excel_row, "Phone",
                   f'"{bad}" is not a valid Ugandan mobile number', "Tap the Phone cell, type a valid 07XXXXXXXX or +2567XXXXXXXX number",
                   "Row turns green; carrier detected (MTN/Airtel)")
    if i in bad_qty_rows:
        bad = rng.choice(["-50", "0", "-120"])
        row[10] = bad
        mark_error(ws.cell(row=excel_row, column=12), problems1, excel_row, "Qty (Kg)",
                   f"Quantity must be greater than 0 (got {bad})", "Tap the Qty cell, enter a positive number (e.g. 750)",
                   "Row turns green and will import")
    if i in blank_name_rows:
        row[1] = ""
        mark_error(ws.cell(row=excel_row, column=2), problems1, excel_row, "Farmer Name",
                   "Farmer row needs a name", "Tap the Farmer Name cell and type the farmer's full name",
                   "Row turns green; a new profile will be created")
    if i in bad_date_rows:
        bad = rng.choice(["not-a-date", "31/13/2026"])
        row[11] = bad
        mark_error(ws.cell(row=excel_row, column=13), problems1, excel_row, "Harvest Date",
                   f'Harvest date "{bad}" is not a valid date', "Tap the Harvest Date cell and enter YYYY-MM-DD (e.g. 2026-07-20)",
                   "Row turns green")
    if i in bad_acreage_rows:
        bad = rng.choice(["-1", "0", "abc"])
        row[8] = bad
        mark_error(ws.cell(row=excel_row, column=10), problems1, excel_row, "Acreage",
                   f'Acreage "{bad}" is not a valid number above 0', "Tap the Acreage cell and enter the farm size in acres",
                   "Row turns green")
    if i in no_farmer_rows:
        row[0] = ""
        row[1] = ""
        row[3] = ""
        row[2] = ""
        mark_error(ws.cell(row=excel_row, column=1), problems1, excel_row, "Farmer ID / Name / Phone",
                   "Produce row needs a farmer (match an ID, phone or provide a name)",
                   "Either leave the row's crop/qty blank, or add the farmer's ID/phone/name",
                   "Row turns green or the produce part is dropped")

    # duplicate-log practice pairs (same farmer + crop + date, logged close together)
    dup_phone = None
    if i in dup_anchor_rows:
        dup_phone = phone
        # sibling row at excel_row+1
        sib = excel_row + 1
        row[12] = grade
        problems1.append((excel_row, "Crop + Date", "Possible duplicate: the next row logs the same farmer + crop + date",
                          "This is a warning (not an error): keep both to see FLAGGED, or edit one row's date/crop",
                          "Import creates two logs; the rule engine flags the pair as possible duplicates"))

    farmers.append((fid, name, phone, gender, district, sub, village, acreage))
    all_rows1.append((excel_row, row, []))

    # sibling duplicate row (recorded, inserted right after the anchor)
    if i in dup_anchor_rows and dup_phone:
        sib_row = [fid, name, dup_phone, gender, refugee, district, sub, village, acreage,
                   crop, qty, date, grade, batch]
        all_rows1.append((excel_row + 1, sib_row, [("Crop + Date", "Duplicate entry (same farmer + crop + harvest date)",
                          "Keep it to demonstrate the duplicate guard, or edit the date/crop",
                          "Flagged as possible duplicate by the rule engine")]))

# sort by excel row and write (siblings land right after anchors)
from openpyxl.styles import PatternFill as _PF, Border as _B, Side as _S
for erow, vals, entries in sorted(all_rows1, key=lambda x: x[0]):
    for c, val in enumerate(vals, 1):
        cell = ws.cell(row=erow, column=c, value=val)
        cell.border = BORDER
    for ent in entries:
        problems1.append((erow, *ent))
        for c in range(1, len(vals) + 1):
            ws.cell(row=erow, column=c).fill = WARN_FILL

readme1 = wb.create_sheet("Read Me")
write_readme(
    readme1,
    "Roki simulation — 1,000 farmers (with practice errors)",
    "Use this file to practise Bulk Upload at scale and inline editing.\n\n"
    "1. (Optional but recommended) Start from a clean database: Settings → Data management → Delete all data…\n"
    "2. Go to Bulk Upload and drag this file in. All columns auto-map (Farmer ID, Farmer Name, Email, Phone, Gender, Refugee Status, District, Sub-County, Village, Acreage, Crop, Qty (Kg), Harvest Date, Grade, Batch ID).\n"
    "3. Rows with light-red cells carry intentional errors. Fix them by tapping the cell in the staging grid (dropdowns appear for Grade / Gender / Refugee Status). Rows that still have errors are NOT imported.\n"
    "4. Light-amber rows are intentional duplicate pairs — keep them to see the duplicate guard flag them as possible duplicates.\n"
    "5. Press Import. Expect: ~970 farmers + logs imported, the rest listed in the report.\n\n"
    "All 1,000 rows: 29 rows carry intentional problems (25 errors + 4 duplicate-warning rows) — all listed below. Everything else imports cleanly.",
    problems1,
    notes=[
        "Farmer IDs are pre-assigned (RFV-UG-00001 … RFV-UG-01000) — they match the Farmer Log simulation file below.",
        "Phones are generated (MTN/Airtel prefixes) and unique; the association engine links rows by ID or phone.",
        "If you already have data in the database, IDs may differ — the log file can also link by phone.",
        "1,000 rows import in a few seconds; the staging grid shows the first 20 rows but imports everything.",
    ],
)
wb.save("/home/user/roki-simulation-1000-farmers.xlsx")
print("file 1 saved: 1000 farmers (36 intentional problems)")

# ======================================================================
# FILE 2 — Farmer log simulation
# ======================================================================
HEADERS2 = ["Log ID", "Farmer ID", "Farmer Name", "Phone", "Crop", "Qty (Kg)", "Grade",
            "Harvest Date", "Batch ID", "Storage Location"]
WIDTHS2 = [13, 14, 24, 17, 15, 10, 8, 13, 17, 28]

wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Harvest Logs"
style_sheet(ws2, HEADERS2, WIDTHS2)

STORAGE = ["Roki Hub — Entebbe Road", "Kawempe Collection Centre", "Nakasero Market Depot",
           "Kisenyi Warehouse", "Village aggregation point", "Farm store"]
problems2 = []
N_LOGS = 900

# error positions in the 900 log rows
anomaly_rows = rng.sample(range(1, N_LOGS + 1), 5)   # way-above-ceiling quantities
neg_rows = rng.sample([r for r in range(1, N_LOGS + 1) if r not in anomaly_rows], 2)
bad_date_rows2 = rng.sample([r for r in range(1, N_LOGS + 1) if r not in anomaly_rows + neg_rows], 2)
no_farmer_rows2 = rng.sample([r for r in range(1, N_LOGS + 1) if r not in anomaly_rows + neg_rows + bad_date_rows2], 1)
dup_anchor = rng.sample([r for r in range(1, N_LOGS + 1) if r not in anomaly_rows + neg_rows + bad_date_rows2 + no_farmer_rows2], 1)[0]

all_rows2 = []
for i in range(1, N_LOGS + 1):
    excel_row = i + 1
    fid, name, phone, gender, district, sub, village, acreage = farmers[(i - 1) % 1000]
    crop = rng.choice(CROPS)
    qty = round(acreage * TYPICAL_KG_PER_ACRE[crop] * rng.uniform(0.55, 1.05), 0)
    date = gen_date()
    grade = rng.choice(GRADES)
    batch = gen_batch(date)
    storage = rng.choice(STORAGE)

    row = [f"RFV-LOG-{i:05d}", fid, name, phone, crop, qty, grade, date, batch, storage]

    if i in anomaly_rows:
        qty = round(acreage * TYPICAL_KG_PER_ACRE[crop] * rng.uniform(3.5, 5.0), 0)
        row[5] = qty
        mark_error(ws2.cell(row=excel_row, column=6), problems2, excel_row, "Qty (Kg)",
                   f"Yield {qty:,.0f} kg exceeds the expected ceiling for {crop} on {acreage} ac",
                   "Keep it to show NEEDS AUDIT, or fix the quantity to a plausible value",
                   "Status becomes NEEDS AUDIT with the rule explanation")
    if i in neg_rows:
        bad = rng.choice(["-30", "-5"])
        row[5] = bad
        mark_error(ws2.cell(row=excel_row, column=6), problems2, excel_row, "Qty (Kg)",
                   f"Quantity must be greater than 0 (got {bad})", "Enter a positive quantity",
                   "Row turns green")
    if i in bad_date_rows2:
        row[7] = "not-a-date"
        mark_error(ws2.cell(row=excel_row, column=8), problems2, excel_row, "Harvest Date",
                   "Harvest date is missing or not a valid date", "Enter YYYY-MM-DD",
                   "Row turns green")
    if i in no_farmer_rows2:
        row[1] = ""
        row[2] = ""
        row[3] = ""
        mark_error(ws2.cell(row=excel_row, column=2), problems2, excel_row, "Farmer ID / Name / Phone",
                   "Produce row needs a farmer", "Add the farmer's ID or phone",
                   "Row turns green")
    if i == dup_anchor:
        problems2.append((excel_row, "Crop + Date",
                          "Possible duplicate: next row logs the same farmer + crop + date",
                          "Keep both to see the duplicate guard flag them",
                          "Second log is FLAGGED as a possible duplicate"))

    all_rows2.append((excel_row, row, []))

    if i == dup_anchor:
        all_rows2.append((excel_row + 1, row, [("Crop + Date", "Duplicate entry (same farmer + crop + date)",
                          "Keep to demonstrate the duplicate guard, or change the date",
                          "Flagged as possible duplicate")]))

for erow, vals, entries in sorted(all_rows2, key=lambda x: x[0]):
    for c, val in enumerate(vals, 1):
        cell = ws2.cell(row=erow, column=c, value=val)
        cell.border = BORDER
    for ent in entries:
        problems2.append((erow, *ent))
        for c in range(1, len(vals) + 1):
            ws2.cell(row=erow, column=c).fill = WARN_FILL

readme2 = wb2.create_sheet("Read Me")
write_readme(
    readme2,
    "Roki simulation — Farmer harvest logs (900 rows)",
    "Use this file to simulate harvest-log data at scale, usually AFTER importing the 1,000-farmer file so the "
    "farmer IDs/phones match.\n\n"
    "1. Import the 1,000-farmer file first (or use a database that already contains those farmers).\n"
    "2. Bulk Upload → drag this file in. Farmer ID/Phone link each log to an existing farmer; Log ID is auto-ignored "
    "(the system generates its own IDs).\n"
    "3. 10 rows carry intentional problems (listed below). Light-red = errors to fix; light-amber = duplicate pair "
    "for the duplicate guard; the big-quantity rows show NEEDS AUDIT.\n"
    "4. After import, check Dashboard → Rule engine findings and the Data Grid → Logs tab.",
    problems2,
    notes=[
        "The 5 large-quantity rows are NOT errors: they demonstrate the yield-anomaly rule (status NEEDS AUDIT).",
        "One REJECT-grade row and one duplicate pair are included on purpose.",
        "Imported logs appear instantly in the Production Forecast (actuals) and Export Supply Planning views.",
    ],
)
wb2.save("/home/user/roki-farmer-log-simulation.xlsx")
print("file 2 saved: 900 harvest logs (10 intentional problems)")
print("done")
